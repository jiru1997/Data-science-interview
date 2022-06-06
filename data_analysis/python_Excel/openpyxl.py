#------------------------------------------------------------------------------------------------------------------------------------#
#workbook / sheet
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

wb = load_workbook('path')
wb.sheetnames

for sheet in wb:
	print(sheet.title)

sheet1 = wb.active
sheet1 = wb['Sheet1']
sheet1 = wb.get_sheet_by_name('Sheet1')
sheet2 = wb.create_sheet('newsheet')

sheet1.max_row / sheet1.max_column

get_column_letter(2)
column_index_from_string('A')

#------------------------------------------------------------------------------------------------------------------------------------#
#cell
cell = sheet['A1']
cell.row / cell.column / cell.value

sheet.cell(1, 1) #first index 1

#------------------------------------------------------------------------------------------------------------------------------------#
#rows / columns
col1 = sheet1['C']
col1 = sheet1['B:C']
col1[2].value



wb.save('path')
